import argparse
import openpyxl

DEFAULT_INPUT = ['openapi.properties', 'openapi.en.properties']
DEFAULT_OUTPUT = 'openapi.xlsx'


def _fetch_file(path: str) -> dict:
    kv = dict()
    with open(path, 'r') as f:
        for line in f.readlines():
            if line.startswith('#'):
                continue
            parts = line.split('=', 1)
            if not parts:
                continue
            key = parts[0]
            if not key:
                continue
            value = parts[1].strip() if len(parts) > 1 else ''
            if key in kv:
                raise KeyError(f"Duplicated key '{key}'")
            kv[key] = value
    return kv


def _join(left: dict, right: dict = None) -> dict:
    merged = dict()
    for k, lv in left.items():
        wrapped = lv if isinstance(lv, list) else [lv]
        if right:
            rv = right.get(k)
            wrapped.append(rv if rv else '')
        merged[k] = wrapped
    return merged


def _to_xlsx(outfile: str, merged: dict):
    file = openpyxl.Workbook()
    sheet = file.active
    r = 1
    for k, wrapped in merged.items():
        sheet.cell(r, 1, k)
        for c in range(len(wrapped)):
            sheet.cell(r, c + 2, wrapped[c])
        r += 1
    file.save(outfile)


def prop_to_xlsx(infiles: list, outfile: str):
    properties = None
    for infile in infiles:
        if not properties:
            properties = _fetch_file(infile)
            continue
        properties = _join(properties, _fetch_file(infile))

    _to_xlsx(outfile, properties)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input', action='extend', nargs='+', default=DEFAULT_INPUT)
    parser.add_argument('-o', '--output', default=DEFAULT_OUTPUT)

    args = parser.parse_args()
    prop_to_xlsx(args.input, args.output)
